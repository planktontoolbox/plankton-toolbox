#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
# Copyright (c) 2010-2016 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).
#
from __future__ import unicode_literals

import os
import shutil
import math
import openpyxl
import PyQt4.QtCore as QtCore
import toolbox_utils
import plankton_core

@toolbox_utils.singleton
class PlanktonCounterManager(QtCore.QObject):
    """ """
    def __init__(self,
                 dataset_dir_path = 'plankton_toolbox_counter/datasets',
                 ):
        """ """
        QtCore.QObject.__init__(self)
        #
        self._dataset_dir_path = dataset_dir_path
        # Check if exists. Create if not.
        if (self._dataset_dir_path) and (not os.path.exists(self._dataset_dir_path)):
            try:
                os.makedirs(self._dataset_dir_path)
#                 print('Directories created for this path: ' + self._dataset_dir_path)
            except Exception as e:
                raise UserWarning('Can\'t create directories in path. Path: ' + self._dataset_dir_path + '. Exception: ' + e)
        #

    def _emit_change_notification(self):
        """ Emit signal to update GUI lists for available datasets and samples. """
        self.emit(QtCore.SIGNAL('planktonCounterListChanged'))
    
    # === Datasets ===
    
    def get_dataset_dir_path(self):
        """ """
        return self._dataset_dir_path
    
    def get_dataset_names(self):
        """ Returns a list with available datasets. """
        datasetnames = []
        if (self._dataset_dir_path) and (os.path.exists(self._dataset_dir_path)):
            for datasetdir in os.listdir(self._dataset_dir_path):
                if os.path.isdir(os.path.join(self._dataset_dir_path, datasetdir)):
                    datasetnames.append(datasetdir)
            return sorted(datasetnames)
        else:
            raise UserWarning('The directory ' + self._dataset_dir_path + ' does not exists.')

    def create_dataset(self, dataset_name):
        """ Creates a new dataset (= a new directory). """
        path = os.path.join(self._dataset_dir_path, dataset_name)
        # Check if exists.
        if (path) and (not os.path.exists(path)):
            try:
                os.makedirs(path)
#                 print('Directories created for this path: ' + path)
            except Exception as e:
                raise UserWarning('Can\'t create directories in path. Path: ' + path + '. Exception: ' + e)
        else:
            raise UserWarning('Dataset already exists, create failed. Dataset name: ' + dataset_name)
        #
        self._emit_change_notification()

    def delete_dataset(self, dataset_name):
        """ Deletes a dataset (the directory and all its content.) """
        path = os.path.join(self._dataset_dir_path, dataset_name)
        # Check if exists.
        if path and os.path.exists(path):
            try:
                shutil.rmtree(path)
            except Exception as e:
                raise UserWarning('Can\'t delete directory. Path: ' + path + '. Exception: ' + e)
        else:
            raise UserWarning('Dataset did not exist, delete failed. Dataset name: ' + dataset_name)
        #
        self._emit_change_notification()

    # === Samples ===
    
    def get_sample_names(self, dataset_name):
        """ Returns a list with available samples within a dataset. """
        samplenames = []
        path = os.path.join(self._dataset_dir_path, dataset_name)
        # Check if exists.
        if (path) and (os.path.exists(path)):
            for sampledir in os.listdir(path):
                if os.path.isdir(os.path.join(path, sampledir)):
                    samplenames.append(sampledir)
            return sorted(samplenames)
        else:
            raise UserWarning('Dataset does not exist. Dataset name: ' + dataset_name)

    def create_sample(self, dataset_name, sample_name):
        """ Creates a new sample (= a new directory in the dataset directory). """
        if not dataset_name:
            dataset_name = 'Default-dataset'
        if not sample_name:
            raise UserWarning('Can\'t create sample. Sample name is missing.')
        #
        path = os.path.join(self._dataset_dir_path, dataset_name, sample_name)
        # Check if exists.
        if (path) and (not os.path.exists(path)):
            try:
                os.makedirs(path)
#                 print('Directories created for this path: ' + path)
            except Exception as e:
                raise UserWarning('Can\'t create directories in path. Path: ' + path + '. Exception: ' + e)
        else:
            raise UserWarning('Sample already exists, create failed. Dataset name: ' + dataset_name)
        #
        self._emit_change_notification()

    def delete_sample(self, dataset_name, sample_name):
        """ Deletes a sample (the directory and all its content.) """
        path = os.path.join(self._dataset_dir_path, dataset_name, sample_name)
        # Check if exists.
        if path and os.path.exists(path):
            try:
                shutil.rmtree(path)
            except Exception as e:
                raise UserWarning('Can\'t delete sample. Path: ' + path + '. Exception: ' + e)
        else:
            raise UserWarning('Sample does not exist, delete failed. Sample name: ' + sample_name)
        #
        self._emit_change_notification()

    def rename_sample(self, dataset_name, old_sample_name, new_sample_name):
        """ Renames a sample. """
        path = os.path.join(self._dataset_dir_path, dataset_name, old_sample_name)
        new_path = os.path.join(self._dataset_dir_path, dataset_name, new_sample_name)
        # Check if exists.
        if path and os.path.exists(path):
            try:
                shutil.move(path, new_path)
            except Exception as e:
                raise UserWarning('Can\'t rename sample. Path: ' + path + '. Exception: ' + e)
        else:
            raise UserWarning('Sample does not exist, rename failed. Dataset name: ' + old_sample_name)
        #
        self._emit_change_notification()

    # === Dataset metadata ===
    
    def load_dataset_metadata(self, dataset_name):
        """ """
        path = os.path.join(self._dataset_dir_path, dataset_name)
        tablefilereader = toolbox_utils.TableFileReader(
                    file_path = path,
                    text_file_name = 'dataset_metadata.txt',                 
                    )
        # Merge header and rows.
        self._current_dataset_metadata = [tablefilereader.header()] + tablefilereader.rows()
#         print('Dataset metadata: ' + unicode(self._current_dataset_metadata))

    def write_dataset_metadata(self, dataset_name):
        """ """
        raise UserWarning('Method: write_dataset_metadata(). Not implemented yet.')

#     # === Sampling info ===
#     
#     def load_sampling_info(self, dataset_name):
#         """ """
#         path = os.path.join(self._dataset_dir_path, dataset_name)
#         tablefilereader = toolbox_utils.TableFileReader(
#                     file_path = path,
#                     text_file_name = 'sampling_info.txt',               
#                     )
#         # Merge header and rows.
#         self._current_sampling_info = [tablefilereader.header()] + tablefilereader.rows()
# #         print('Sampling info: ' + unicode(self._current_sampling_info))

    # === Import/Export ===
    
    def import_dataset(self, source_path, source_file_name, dataset_name):
        """ """
        raise UserWarning('Method: import_dataset(). Not implemented yet.')

    def export_dataset(self, dataset_name, target_path, target_file_name):
        """ """
        raise UserWarning('Method: export_dataset(). Not implemented yet.')

    def import_orderer_info(self, source_path, source_file_name, dataset_name):
        """ """
        raise UserWarning('Method: import_orderer_info(). Not implemented yet.')

    def import_sampling_info(self, source_path, source_file_name, dataset_name):
        """ """
        raise UserWarning('Method: import_sampling_info(). Not implemented yet.')


class PlanktonCounterSample():
    """ Manager for rows in sample. """
    def __init__(self,
                 dataset_dir_path = None,
                 dataset_name = None,
                 sample_name = None,
                 ):
        """ """
        self._dataset_dir_path = dataset_dir_path
        self._dataset_name = dataset_name
        self._sample_name = sample_name
        #
        self._sample_info_dict = {} # <key>: <value>
        self._sample_rows = {} # <row_key>: <SampleRow-object>
        self._sample_header = ['scientific_full_name', 
                               'taxon_class', 
                               'scientific_name', 
                               'size_class', 
                               'method_step',
                               'count_area_number',
                               'locked_at_area',
                               'counted_units', 
                               'counted_units_list', 
                               'abundance_class', 
                               'coefficient', 
                               'abundance_units_l', 
                               'volume_mm3_l', 
                               'carbon_ugc_l', 
                               'volume_um3_unit', 
                               'carbon_pgc_unit',
                               'variable_comment',
                               'trophic_type', 
                               'unit_type', 
                               'species_flag_code',
                               'cf',
                               ]

        # Create file writers.
        if not self._dataset_name:
            self._dataset_name = 'Default-dataset'
        #
        path = os.path.join(self._dataset_dir_path, self._dataset_name, self._sample_name)
        self._tablefilewriter_sample_data = toolbox_utils.TableFileWriter(
            file_path = path,
            text_file_name = 'sample_data.txt',                 
            )
        self._tablefilewriter_sample_info = toolbox_utils.TableFileWriter(
            file_path = path,
            text_file_name = 'sample_info.txt',                 
            )
        # Load from file.
        self.load_sample_info()
        self.load_sample_data()
        
    def get_dir_path(self):
        """ """
        return os.path.join(self._dataset_dir_path, self._dataset_name, self._sample_name)
    
    def clear(self):
        """ """
        self._sample_rows = {} # <row_key>: <CounterRow-object>
        
    def get_sample_info(self):
        """ """
        return self._sample_info_dict

    def set_sample_info(self, sample_info_dict):
        """ """
        self._sample_info_dict = sample_info_dict

    def save_sample_info(self):
        """ """
        header = ['key', 'value']
        rows = []
        for key in sorted(self._sample_info_dict.keys()):
            rows.append([key, self._sample_info_dict[key]])
        #
        self._tablefilewriter_sample_info.write_file(header, rows)

    def update_sample_info(self, info_dict):
        """ """
        for key in info_dict.keys():
            self._sample_info_dict[key] = info_dict[key]

    def load_sample_info(self):
        """ """
        self._sample_info_dict = {}
        #
        if (self._dataset_dir_path is None) or (self._dataset_name is None) or (self._sample_name is None):
            raise UserWarning('Failed to load sample file. Path, dataset name or sample name is missing.')
        # Create file if not exists.
        path = os.path.join(self._dataset_dir_path, self._dataset_name, self._sample_name)
        if not os.path.exists(os.path.join(path, 'sample_info.txt')):
            header = ['key', 'value']
            self._tablefilewriter_sample_info.write_file(header, [])
        # Read file to dict.
        tablefilereader = toolbox_utils.TableFileReader(
                    file_path = path,
                    text_file_name = 'sample_info.txt',                 
                    )
        for row in tablefilereader.rows():
            if len(row) >= 2:
                key = row[0].strip()
                value = row[1].strip()
                self._sample_info_dict[key] = value

    def get_sample_header_and_rows(self):
        """ """
        header = self.get_header()
        rows = self.get_rows()
        return header, rows
 
    def get_header(self):
        """ """
        return self._sample_header

    def get_rows(self):
        """ """
        rows = []
        for key in sorted(self._sample_rows.keys()):
            rows.append(self._sample_rows[key].get_row_as_text_list(self._sample_header))
        return rows

    def load_sample_data(self):
        """ """
        self._sample_rows = {}
        #
        if (self._dataset_dir_path is None) or (self._dataset_name is None) or (self._sample_name is None):
            raise UserWarning('Failed to load sample file. Path, dataset name or sample name is missing.')
        # Create file if not exists.
        path = os.path.join(self._dataset_dir_path, self._dataset_name, self._sample_name)
        if not os.path.exists(os.path.join(path, 'sample_data.txt')):
            self._tablefilewriter_sample_data.write_file( self._sample_header, [])
        # Read sample data to self._sample_rows.
        tablefilereader = toolbox_utils.TableFileReader(
                    file_path = path,
                    text_file_name = 'sample_data.txt',                 
                    )
        self.update_all_sample_rows(tablefilereader.header(), tablefilereader.rows())
        
    def update_all_sample_rows(self, header, rows):
        """ """
        self._sample_rows = {}
        for row in rows:
            if (len(row) >= 2) and (len(row[0]) >= 0) and (len(row[1]) >= 0):
                info_dict = dict(zip(header, row))
                sample_row = SampleRow(info_dict)
                # Don't save samples without name or counted value = 0.
                if len(sample_row.get_scientific_name()) > 0:
                    if sample_row.get_counted_units() > 0:
                        self._sample_rows[sample_row.get_key()] = sample_row

    def recalculate_coefficient(self, current_method):
        """ """
        # Recalculate all rows.
        for sampleobject in self._sample_rows.values():            
            # Get coefficient from method.
            method_step = sampleobject.get_method_step()
            method_step_fields = current_method.get_counting_method_step_fields(method_step)
            coefficient_one_unit = int(method_step_fields.get('coefficient_one_unit', '0'))
            # Calculate coefficient for one row.
            count_area_number = int(sampleobject.get_count_area_number())
            locked_at_area = sampleobject.get_locked_at_area()
            ###counted_units_list = sampleobject.get_counted_units_list()
            #
            number_of_areas = count_area_number
            if locked_at_area:
                number_of_areas = int(locked_at_area)
            #
            coefficient = int((coefficient_one_unit / number_of_areas) + 0.5)
            sampleobject.set_coefficient(unicode(coefficient))
            # 
            sampleobject._calculate_values()

    def save_sample_data(self):
        """ """
        rows = []
        for key in sorted(self._sample_rows.keys()):
            rows.append(['\t'.join(self._sample_rows[key].get_row_as_text_list(self._sample_header))])
        #
        self._tablefilewriter_sample_data.write_file(self._sample_header, rows)
        
    def get_taxa_summary(self, summary_type = None, 
                               most_counted_sorting = False, 
                               method_step = None):
        """ """
        summary_data = []
        #
        totalcounted = 0
        countedspecies = {} # Value for sizeclasses aggregated.
        size_range_dict = {} # Value for sizeclasses aggregated.
        locked_list = []
        #
        for sampleobject in self._sample_rows.values():
            # Check method step.
            if method_step:
                if not method_step == sampleobject.get_method_step():
                    continue
            # Count on scientific name. Standard alternative.
            taxon = sampleobject.get_scientific_full_name()
            size = sampleobject.get_size_class()
            # Use the same key for locked items.
            if sampleobject.is_locked():
                if size:
                    locked_list.append(taxon + ' [' + size + '] ')     
                else:
                    locked_list.append(taxon)     
            # Count on class name.
            if summary_type == 'Counted per classes':
                taxon = plankton_core.Species().get_taxon_value(taxon, 'taxon_class')
                if len(taxon) == 0:
                    taxon = '<class unknown>'
            # Count on scientific name and size class.
            elif summary_type == 'Counted per taxa':
                if size:
                    taxon = taxon
            elif summary_type == 'Counted per taxa/sizes':
                if size:
                    taxon = taxon + ' [' + size + '] '
            # Create in list, first time only.
            if taxon not in countedspecies:
                countedspecies[taxon] = 0
            # Add.
            try:
                abundance_class = sampleobject.get_abundance_class()
                if abundance_class in ['', '0']:
                    countedspecies[taxon] += int(sampleobject.get_counted_units())
                    totalcounted += int(sampleobject.get_counted_units())
                else:
                    if summary_type in ['Counted per taxa', 'Counted per taxa/sizes']:
                        if abundance_class == '1':
                            countedspecies[taxon] = '1=Observed'
                        elif abundance_class == '2':
                            countedspecies[taxon] = '2=Several cells'
                        elif abundance_class == '3':
                            countedspecies[taxon] = '3=1-10%'
                        elif abundance_class == '4':
                            countedspecies[taxon] = '4=10-50%'
                        elif abundance_class == '5':
                            countedspecies[taxon] = '5=50-100%'
                    else:
                        countedspecies[taxon] = 'Qualitative'
            except: pass # If value = ''.
            #
            if summary_type in ['Counted per taxa/sizes']:
                bvol_size_range = plankton_core.Species().get_bvol_dict(sampleobject.get_scientific_name(), size).get('bvol_size_range', '')
                if bvol_size_range:
                    size_range_dict[taxon] = '(Size: ' + bvol_size_range + ')'
        #    
        summary_data.append('Total counted: ' + unicode(totalcounted))
        summary_data.append('')
        if most_counted_sorting == False:
            # Alphabetical.
            for key in sorted(countedspecies):
                size_range = ''
                if key in size_range_dict:
                    size_range = '    ' + size_range_dict[key]
                lock_info = ''
                if key in locked_list:
                    lock_info = ' [Locked]'
                
                summary_data.append(key + ': ' + unicode(countedspecies[key]) + lock_info + size_range)
        else:
            # Sort for most counted.
            for key in sorted(countedspecies, key=countedspecies.get, reverse=True):
                size_range = ''
                if key in size_range_dict:
                    size_range = ' ' + size_range_dict[key]
                lock_info = ''
                if key in locked_list:
                    lock_info = ' [Locked]'
                summary_data.append(key + ': ' + unicode(countedspecies[key]) + lock_info + size_range)
        #
        return summary_data

    def get_locked_taxa(self, method_step = None):
        """ """
        species_locked_list = []
        #
        for sampleobject in self._sample_rows.values():
            # Check method step.
            if method_step:
                if not method_step == sampleobject.get_method_step():
                    continue
            # 
            taxon = sampleobject.get_scientific_full_name()
            size = sampleobject.get_size_class()
            # 
            if sampleobject.is_locked():
                species_locked_list.append([taxon, size, True])
            else:
                species_locked_list.append([taxon, size, False])   
        #
        return species_locked_list
    
    def lock_taxa(self, scientific_full_name, size_class, locked_at_count_area):
        """ """
        search_dict = {}
        search_dict['scientific_full_name'] = scientific_full_name
        search_dict['size_class'] = size_class
        samplerowkey = SampleRow(search_dict).get_key()
        if samplerowkey in self._sample_rows:
            if not self._sample_rows[samplerowkey].is_locked():
                self._sample_rows[samplerowkey].set_lock(locked_at_count_area)
    
    def unlock_taxa(self, scientific_full_name, size_class, count_area_number):
        """ """
        search_dict = {}
        search_dict['scientific_full_name'] = scientific_full_name
        search_dict['size_class'] = size_class
        samplerowkey = SampleRow(search_dict).get_key()
        if samplerowkey in self._sample_rows:
            self._sample_rows[samplerowkey].set_lock('')
        # 
        self._sample_rows[samplerowkey].set_count_area_number(count_area_number)
    
    def get_sample_row_dict(self, counted_row_dict):
        """ """
        samplerowkey = SampleRow(counted_row_dict).get_key()
        if samplerowkey in self._sample_rows:
            return self._sample_rows[samplerowkey].get_sample_row_dict()
        #
        return {}
        
    def update_sample_row(self, counted_row_dict):
        """ """
        if len(counted_row_dict.get('scientific_name', '')) > 0:
            samplerowkey = SampleRow(counted_row_dict).get_key()
            if samplerowkey in self._sample_rows:
                self._sample_rows[samplerowkey].update_sample_row_dict(counted_row_dict)

    def get_counted_value(self, selected_dict):
        """ """
        samplerowkey = SampleRow(selected_dict).get_key()
        if samplerowkey in self._sample_rows:
            return self._sample_rows[samplerowkey].get_counted_units()
        else:
            return 0
        
    def update_counted_value_in_core(self, counted_row_dict, value):
        """ """
        if value == '0':
            # Delete row.
            samplerowkey = SampleRow(counted_row_dict).get_key()
            if samplerowkey in self._sample_rows:
                del self._sample_rows[samplerowkey]
            return
        #
        if len(counted_row_dict.get('scientific_full_name', '')) > 0:
            samplerowkey = SampleRow(counted_row_dict).get_key()
            if samplerowkey not in self._sample_rows:
                self._sample_rows[samplerowkey] = SampleRow(counted_row_dict)
            # Check if the same method step or locked taxa.
            samplerowobject = self._sample_rows[samplerowkey]
            # Don't check for validity when the value is same same.
            if samplerowobject.get_counted_units() == value:
                return
            if  samplerowobject.is_locked():
                raise UserWarning('Selected taxon is locked') 

            
#             if counted_row_dict.get('method_step') == samplerowobject.get_method_step():
            if True:
                samplerowobject.set_counted_units(value)
                samplerowobject.update_sample_row_dict(counted_row_dict)
            else:
                raise UserWarning('Selected taxon is already counted in another method step.')

    def update_abundance_class_in_core(self, counted_row_dict, value):
        """ """
        if value == '0':
            # Delete row.
            samplerowkey = SampleRow(counted_row_dict).get_key()
            if samplerowkey in self._sample_rows:
                del self._sample_rows[samplerowkey]
            return
        #
        if len(counted_row_dict.get('scientific_full_name', '')) > 0:
            samplerowkey = SampleRow(counted_row_dict).get_key()
            if samplerowkey not in self._sample_rows:
                self._sample_rows[samplerowkey] = SampleRow(counted_row_dict)
            # Check if the same method step or locked taxa.
            samplerowobject = self._sample_rows[samplerowkey]
            # Don't check for validity when the value is same same.
            if samplerowobject.get_counted_units() == value:
                return
            if  samplerowobject.is_locked():
                raise UserWarning('Selected taxon is locked') 

            
#             if counted_row_dict.get('method_step') == samplerowobject.get_method_step():
            if True:
#                 samplerowobject.set_counted_units(value)
                samplerowobject.set_abundance_class(value)
                samplerowobject.update_sample_row_dict(counted_row_dict)
            else:
                raise UserWarning('Selected taxon is already counted in another method step.')
    
    def delete_rows_in_method_step(self, current_method_step):
        """ """
        for sampleobject in self._sample_rows.values():
            if sampleobject.get_method_step() == current_method_step:
                del self._sample_rows[sampleobject.get_key()]
    
    def update_coeff_for_sample_rows(self, current_method_step, count_area_number, coefficient):
        """ """
        for sampleobject in self._sample_rows.values():
            if sampleobject.get_method_step() == current_method_step:
                if not sampleobject.is_locked():
                    sampleobject.set_count_area_number(count_area_number)
                    sampleobject.set_coefficient(coefficient)
    
    def import_sample_from_excel(self, excel_file_path):
        """ Import from Excel. """
        # Sample info.
        tablefilereader = toolbox_utils.TableFileReader(
                    file_path = '',
                    excel_file_name = excel_file_path,
                    excel_sheet_name = 'sample_info.txt',                 
                    )
        sample_header = tablefilereader.header()
        sample_rows = tablefilereader.rows()
        #
        self._tablefilewriter_sample_info.write_file(sample_header, sample_rows)

        # Sample data.
        tablefilereader = toolbox_utils.TableFileReader(
                    file_path = '',
                    excel_file_name = excel_file_path,
                    excel_sheet_name = 'sample_data.txt',                 
                    )
        data_header = tablefilereader.header()
        data_rows = tablefilereader.rows()
        #
        self._tablefilewriter_sample_data.write_file(data_header, data_rows)
        
        # Sample method.
        tablefilereader = toolbox_utils.TableFileReader(
                    file_path = '',
                    excel_file_name = excel_file_path,
                    excel_sheet_name = 'counting_method.txt',                 
                    )
        method_header = tablefilereader.header()
        method_rows = tablefilereader.rows()
        
        path = os.path.join(self._dataset_dir_path, self._dataset_name, self._sample_name)
        tablefilewriter_sample_method = toolbox_utils.TableFileWriter(
            file_path = path,
            text_file_name = 'counting_method.txt',                 
            )
        #
        tablefilewriter_sample_method.write_file(method_header, method_rows)

    def export_sample_to_excel(self, export_target_dir, export_target_filename):
        """ Export to Excel. """
        # Prepare sample info header and rows.
        sample_info_dict = self.get_sample_info()
        #
        sample_info_header = ['key', 'value']
        sample_info_header_order = [
                'plankton_toolbox_version',
                'sample_name',
                'sample_id',
                'sample_date',
                'sample_time',
                'visit_year',
                'country_code',
                'platform_code',
                'sampling_series',
                'sampling_laboratory',
                'orderer',                
                'project_code',
                'station_name',
                'sample_latitude_dm',
                'sample_longitude_dm',
                'sample_latitude_dd',
                'sample_longitude_dd',
                'sample_min_depth_m',
                'sample_max_depth_m',
                'water_depth_m',
                'sampler_type_code',
                'sampled_volume_l',
                'net_type_code', 
                'sampler_area_m2',
                'net_mesh_size_um',
                'wire_angle_deg',
                'net_tow_length_m',
                'analytical_laboratory', 
                'analysis_date',
                'analysed_by',
                'sample_comment',               
                ]
        sample_info_rows = []
        for header_item in sample_info_header_order:
            sample_info_rows.append([header_item, sample_info_dict.get(header_item, '')])
        
        # Prepare sample info header and rows.
        sample_data_header = self.get_header()
        sample_data_rows = self.get_rows()
        
        # Prepare sample info header and rows.
        sample_method_header = []
        sample_method_rows = []
        sample_path = self.get_dir_path()
        if os.path.exists(os.path.join(sample_path, 'counting_method.txt')):
            sample_method_header, sample_method_rows = plankton_core.PlanktonCounterMethods().get_counting_method_table(
                                                sample_path, 'counting_method.txt')        

        # Use openpyxl for Excel.
        workbook = openpyxl.Workbook(optimized_write = True)  # Supports big files.

#         # Sheet: Summary.
#         sampleinfo_worksheet = workbook.create_sheet('Summary')
#         sampleinfo_worksheet.title = 'Summary'
# #         # Header.
# #         sampleinfo_worksheet.append(sample_info_header)
# #         # Rows.
# #         for row in sample_info_rows:
# #             sampleinfo_worksheet.append(row)
#         sampleinfo_worksheet.append(['This summary worksheet i under development...'])
        
        # Sheet: Table summary.
        sampleinfo_worksheet = workbook.create_sheet('Table summary')
        sampleinfo_worksheet.title = 'Table summary'
        # Header.
        sampleinfo_worksheet.append(sample_info_header_order + sample_data_header)
        # Rows.
        for row in sample_data_rows: 
            row_dict = dict(zip(sample_data_header, row))
            row_dict.update(sample_info_dict)
            table_row = []
            for key in sample_info_header_order + sample_data_header:
                table_row.append(row_dict.get(key, ''))
            #
            sampleinfo_worksheet.append(table_row)
        
        
        # Sheet: Sample info.
        sampleinfo_worksheet = workbook.create_sheet('sample_info.txt')
        sampleinfo_worksheet.title = 'sample_info.txt'
        # Header.
        sampleinfo_worksheet.append(sample_info_header)
        # Rows.
        for row in sample_info_rows:
            sampleinfo_worksheet.append(row)
        # Extra info for restart of counting session.
        sampleinfo_worksheet.append(['']) # Empty.
        for key in sample_info_dict.keys():
            if key == 'last_used_method_step':
                sampleinfo_worksheet.append([key, sample_info_dict.get(key, '')])
            if key.startswith('max_count_area<+>'):
                sampleinfo_worksheet.append([key, sample_info_dict.get(key, '')])
        # Sheet: Sample data.
        sampledata_worksheet = workbook.create_sheet('sample_data.txt')
        sampledata_worksheet.title = 'sample_data.txt'
        # Header.
        sampledata_worksheet.append(sample_data_header)
        # Rows.
        for row in sample_data_rows:
            sampledata_worksheet.append(row)
        # Sheet: Sample method.
        samplemethod_worksheet = workbook.create_sheet('counting_method.txt')
        samplemethod_worksheet.title = 'counting_method.txt'
        # Header.
        samplemethod_worksheet.append(sample_method_header)
        # Rows.
        for row in sample_method_rows:
            samplemethod_worksheet.append(row)
        # Sheet: README.
        samplemethod_worksheet = workbook.create_sheet('README')
        samplemethod_worksheet.title = 'README'
        # Header.
        samplemethod_worksheet.append(['Plankton Toolbox - Plankton counter'])
        # Rows.
        readme_text = [
            [''],
            ['This Excel file is generated by Plankton Toolbox.'],
            [''],
            ['The file represents one counted plankton sample. It can be used for export and import '],
            ['between different computers running Plankton Toolbox, or as an archive file.'],
            ['Don\'t edit or rename the sheets sample_info.txt, sample_data.txt or '],
            ['counting_method.txt if the file should be imported later to Plankton Toolbox.'],
            [''],
            ['Generated sheets:'],
            ['- Summary: Page containing information in a compact format. (Under development, not available yet.)'],
            ['- Table summary: Page with generated data in table format.'],
            ['- sample_info.txt: Copy of the internally used file for information connected to a sample (metadata).'],
            ['- sample_data.txt: Copy of the internally used file for counted sample rows.'],
            ['- counting_method.txt: Copy of the internally used file for parameters related to the counting method.'],
            ['- README: This text.'],
            [''],
            ['More info at: http://nordicmicroalgae.org and http://plankton-toolbox.org '],
            ]
        # 
        for row in readme_text:
            samplemethod_worksheet.append(row)
            
        # Save to file.
        filepathname = os.path.join(export_target_dir, export_target_filename)
        workbook.save(filepathname)


class SampleRow():
    """ Defines the content of one counted sample row. """
    def __init__(self, sample_row_dict):
        """ """
        self._sample_row_dict = {}
        self._sample_row_dict.update(sample_row_dict)
        #
        self._scientific_full_name = self._sample_row_dict.get('scientific_full_name', '')
        self._scientific_name = self._sample_row_dict.get('scientific_name', '')
        self._size_class = self._sample_row_dict.get('size_class', '')
        #
        # Get species related dictionaries for this taxon/sizeclass.
        self._taxon_dict = plankton_core.Species().get_taxon_dict(self._scientific_name)
        self._size_class_dict = plankton_core.Species().get_bvol_dict(self._scientific_name, self._size_class)
        self._sample_row_dict['taxon_class'] = self._taxon_dict.get('taxon_class', '')
        self._sample_row_dict['trophic_type'] = self._size_class_dict.get('bvol_trophic_type', '')
        self._sample_row_dict['unit_type'] = self._size_class_dict.get('bvol_unit', '')
        #
        self._bvol_volume = 0.0
        self._bvol_carbon = 0.0
        try:
            self._bvol_volume = float(self._size_class_dict.get('bvol_calculated_volume_um3', '0').replace(',', '.'))
            self._bvol_carbon = float(self._size_class_dict.get('bvol_calculated_carbon_pg', '0').replace(',', '.'))
        except:
            pass
        self._sample_row_dict['volume_um3_unit'] = unicode(self._round_value(self._bvol_volume))
        self._sample_row_dict['carbon_pgc_unit'] = unicode(self._round_value(self._bvol_carbon))

    def get_sample_row_dict(self):
        """ """
#         print('DEBUG GET scientific_full_name: ' + self._sample_row_dict.get('scientific_full_name', ''))
#         print('DEBUG GET size_class: ' + self._sample_row_dict.get('size_class', ''))
#         print('DEBUG GET variable_comment: ' + self._sample_row_dict.get('variable_comment', ''))
        return self._sample_row_dict

    def update_sample_row_dict(self, sample_row_dict):
        """ """
#         print('DEBUG UPDATE scientific_full_name: ' + sample_row_dict.get('scientific_full_name', ''))
#         print('DEBUG UPDATE size_class: ' + sample_row_dict.get('size_class', ''))
#         print('DEBUG UPDATE variable_comment: ' + sample_row_dict.get('variable_comment', ''))
        self._sample_row_dict.update(sample_row_dict)
    
    def get_key(self):
        """ """
        rowkey = self._scientific_full_name + '+' + self._size_class
        return rowkey
        
    def get_scientific_full_name(self):
        """ """
        return self._scientific_full_name
    
    def get_scientific_name(self):
        """ """
        return self._scientific_name
    
    def get_size_class(self):
        """ """
        return self._size_class
    
    def get_method_step(self):
        """ """
        return self._sample_row_dict.get('method_step', '')
        
    def set_lock(self, locked_at_count_area):
        """ """
        self._sample_row_dict['locked_at_area'] = locked_at_count_area

    def get_locked_at_area(self):
        """ """
        return self._sample_row_dict.get('locked_at_area', '')

    def set_unlock(self):
        """ """
        self._sample_row_dict['locked_at_area'] = ''

    def is_locked(self):
        """ """
        if self._sample_row_dict.get('locked_at_area', '') == '':
            return False
        else:
            return True
    
    def get_count_area_number(self):
        """ """
        return self._sample_row_dict.get('count_area_number', '')
        
    def set_count_area_number(self, count_area_number):
        """ """
        self._sample_row_dict['count_area_number'] = count_area_number
        count_area_number = int(count_area_number)
        # Adjust length of list for counted per are.
        counted_units_list = self._sample_row_dict.get('counted_units_list', None)
        if counted_units_list:
            counted_units_list = [int(x) for x in counted_units_list.split(';')]
        else:
            counted_units_list = count_area_number * [0]
            counted_units_list[0] = self._sample_row_dict.get('counted_units', '0')
        #
        if len(counted_units_list) < count_area_number:
            counted_units_list += (count_area_number - len(counted_units_list)) * [0]
        if len(counted_units_list) > count_area_number:
            counted_units_list = counted_units_list[:count_area_number] 
            # Recalculate when areas are removed.
            calculated_value = sum(counted_units_list)
            self._sample_row_dict['counted_units'] = unicode(calculated_value)
        #
        self._sample_row_dict['counted_units_list'] = ';'.join(unicode(x) for x in counted_units_list)
        # print('DEBUG: add count area: ' + self._sample_row_dict['counted_units_list'])
    
    def set_coefficient(self, coefficient):
        """ """
        self._sample_row_dict['coefficient'] = coefficient

    def get_counted_units_list(self):
        """ """
        counted_units_list = self._sample_row_dict.get('counted_units_list', '0')
        #
        return counted_units_list
    
    def get_counted_units(self):
        """ """
        countedunits = self._sample_row_dict.get('counted_units', '0')
        #
        return countedunits
    
    def set_counted_units(self, value):
        """ """
        old_value = self._sample_row_dict.get('counted_units', 0)
        #
        self._sample_row_dict['counted_units'] = value
        self._sample_row_dict['abundance_class'] = ''
        #
        count_area_number = int(self._sample_row_dict.get('count_area_number', '1'))
        counted_units_list = self._sample_row_dict.get('counted_units_list', None)
        if counted_units_list:
            counted_units_list = [int(x) for x in counted_units_list.split(';')]
        else:
            counted_units_list = count_area_number * [0]
        #    
        if count_area_number == 1:
            counted_units_list[0] = value
        else:
            if old_value != '':
                value_for_area = int(value) - int(old_value) 
            else:
                value_for_area = int(value) 
            counted_units_list[(count_area_number - 1)] += value_for_area
        #
        self._sample_row_dict['counted_units_list'] = ';'.join(unicode(x) for x in counted_units_list)
        #
        #print('DEBUG: ' + self._sample_row_dict['counted_units_list'])
    
    def get_abundance_class(self):
        """ """
        countedunits = self._sample_row_dict.get('abundance_class', '0')
        #
        return countedunits
    
    def set_abundance_class(self, value):
        """ """
        self._sample_row_dict['abundance_class'] = value
        self._sample_row_dict['counted_units'] = ''
    
    def get_row_as_text_list(self, header_list):
        """ """
        self._calculate_values()
        #
        row = [] 
        for header_item in header_list: 
            row.append(self._sample_row_dict.get(header_item, '')) 
        #
        return row 

    def _calculate_values(self):
        """ """
        counted_txt = self._sample_row_dict.get('counted_units', '')
        coefficient_txt = self._sample_row_dict.get('coefficient', '0')
        # Check if abundance_class.
        if counted_txt == '':
            self._sample_row_dict['abundance_units_l'] = ''
            self._sample_row_dict['volume_mm3_l'] = ''
            self._sample_row_dict['carbon_ugc_l'] = ''
            return
        #
        try:
            counted = float(counted_txt.replace(',', '.'))
            coefficient = float(coefficient_txt.replace(',', '.'))
            abundance = counted * coefficient
            self._sample_row_dict['abundance_units_l'] = unicode(self._round_value(abundance))
            #
            try:
                value = abundance * self._bvol_volume / 1000000.0
                self._sample_row_dict['volume_mm3_l'] = unicode(self._round_value(value))
            except:
                self._sample_row_dict['volume_mm3_l'] = '0.00'
            #
            try:
                value = abundance * self._bvol_carbon / 1000.0
                self._sample_row_dict['carbon_ugc_l'] = unicode(self._round_value(value))
            except:
                self._sample_row_dict['carbon_ugc_l'] = '0.00'
        except:
            self._sample_row_dict['abundance_units_l'] = '0.00'
            self._sample_row_dict['volume_mm3_l'] = '0.00'
            self._sample_row_dict['carbon_ugc_l'] = '0.00'
            
    def _round_value(self, value, 
                           n = 4): # Number of significant figures.
        """ """
        if value != 0.0:
            value = round(value, - int(math.floor(math.log10(abs(value)))) + (n - 1)) 
        return value

