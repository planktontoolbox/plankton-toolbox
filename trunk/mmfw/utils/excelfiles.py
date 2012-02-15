#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
# Project: Moray
# Author: Arnold Andreasson, info@mellifica.se
# Copyright (c) 2011 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License as follows:
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import mmfw
try: 
    import openpyxl.workbook as excelworkbook
    import openpyxl.reader.excel as excelreader
    import openpyxl.writer.excel as excelwriter
except ImportError: 
    print("Python package openpyxl missing. Download from http://pypi.python.org/pypi/openpyxl.")
    raise UserWarning("Python package openpyxl missing. Download from http://pypi.python.org/pypi/openpyxl.")

class ExcelFiles():
    """
    This class ...  
    """
    def __init__(self):
        """ """
        self._metadata = {}
        self._header = [] # 
        self._rows = []
        
    def readToTableDataset(self, 
                           target_dataset, 
                           file_name,
                           sheet_name = None, 
                           header_row = 0, 
                           data_rows_from = 1, 
                           data_rows_to = None, # None = read all.
                           used_columns_from = 0, 
                           used_columns_to = None): # None = read all.
        """ """
        if file_name == None:
            raise UserWarning("File name is missing.")
        if not isinstance(target_dataset, mmfw.DatasetTable):
            raise UserWarning("Target dataset is not of valid type.")
        try:
            workbook = excelreader.load_workbook(file_name)
            if workbook == None:
                raise UserWarning("Can't read Excel (.xlsx) file.")
            worksheet = None
            if sheet_name:
                # 
                if sheet_name in workbook.get_sheet_names():
                    worksheet = workbook.get_sheet_by_name(sheet_name)
                else:
                    raise UserWarning("Excel sheet " + sheet_name + " not available.")      
            else:
                # Use the first sheet if not specified.
                worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
            
            # Prepare boundary. 
            if used_columns_to:
                used_columns_to = min(used_columns_to,  worksheet.get_highest_column())
            else:
                used_columns_to = worksheet.get_highest_column()
            if data_rows_to:
                data_rows_to = min(data_rows_to,  worksheet.get_highest_row())
            else:
                data_rows_to = worksheet.get_highest_row()
                
            # Read header row.
            header = []
            for col in xrange(used_columns_from, worksheet.get_highest_column()):
                value = worksheet.cell(row=header_row, column=col).value
                if value:
                    header.append(unicode(value).strip())
                else:
                    header.append(u'')
            target_dataset.setHeader(header)
                    
            # Read data rows.
            for row in xrange(data_rows_from, data_rows_to): 
                newrow = []
                for col in xrange(used_columns_from, worksheet.get_highest_column()):
                    value = worksheet.cell(row=row, column=col).value
                    if value:
                        newrow.append(unicode(value).strip())
                    else:
                        newrow.append(u'')
                target_dataset.appendRow(newrow)
        #  
        except Exception:
            print("Can't read Excel file. File name: " + file_name )
            raise UserWarning("Can't read Excel file. File name: " + file_name)

    def writeTableDataset(self, table_dataset, file_name):
        """ """
        if file_name == None:
            raise UserWarning("File name is missing.")
        if not isinstance(table_dataset, mmfw.DatasetTable):
            raise UserWarning("Dataset is not of a valid type.")
        try:
            workbook = excelworkbook.Workbook() # Create workbook.
            worksheet = workbook.get_active_sheet() # Workbooks contains at least one worksheet.
            # Header.
            for columnindex, item in enumerate(table_dataset.getHeader()):
                cell = worksheet.cell(row=0, column=columnindex)
                cell.value = unicode(item)
            # Rows.
            for rowindex, row in enumerate(table_dataset.getRows()):
                for columnindex, item in enumerate(row):
                    cell = worksheet.cell(row=rowindex + 1, column=columnindex)
                    cell.value = unicode(item)
            #    
            excelwriter.save_workbook(workbook, file_name)
            # Close.
            # TODO: Not needed?        
        except (IOError, OSError):
            mmfw.Logging().log("Failed to write to file: " + file_name)
            raise UserWarning("Failed to write to file: " + file_name)
