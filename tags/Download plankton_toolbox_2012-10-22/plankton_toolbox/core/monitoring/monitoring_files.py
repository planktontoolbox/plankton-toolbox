#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
# Project: Plankton Toolbox. http://plankton-toolbox.org
# Author: Arnold Andreasson, info@mellifica.se
# Copyright (c) 2010-2012 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License as follows:
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

"""


TODO: Not used in current version. Should be rewritten and integrated with envmonlib.


"""


"""
Parses sample files generated by the PW system. File extension should be ".csv".

"""

import codecs
import json
import urllib
import string
import envmonlib
import plankton_toolbox.toolbox.toolbox_settings as toolbox_settings
# Openpyxl is not included in python(xy).
try: 
    import openpyxl.workbook as excelworkbook
    import openpyxl.reader.excel as excelreader
    import openpyxl.writer.excel as excelwriter
except ImportError: 
    print('Python package openpyxl missing.')

class MonitoringFiles(object):
    """ 
    """
    def __init__(self):
        """ """
        self._metadata = {}
        self._header = []
        self._rows = []
        
    def clear(self):
        """ """
        self._metadata = {}
        self._header = []
        self._rows = []

    def getMetadata(self, key):
        """ """
        return self._metadata.get(key, u'')

    def getHeader(self):
        """ """
        return self._header

    def getRows(self):
        """ """
        return self._rows

    def getHeaderItem(self, column):
        """ Used for calls from QAbstractTableModel. """
        try:
            return self._header[column]
        except Exception:
            return ''

    def getDataItem(self, row, column):
        """ Used for calls from QAbstractTableModel. """
        try:
            return self._rows[row][column]
        except Exception:
            return ''

    def getColumnCount(self):
        """ Used for calls from QAbstractTableModel. """
        try:
            return len(self._header)
        except Exception:
            return ''

    def getRowCount(self):
        """ Used for calls from QAbstractTableModel. """
        try:
            return len(self._rows)
        except Exception:
            return ''

#    def SaveAsTextFile(self, fileName):
#        """ """
#        if fileName == None:
#            raise UserWarning('File name is missing.')
#        try:
#            txtencode = toolbox_settings.ToolboxSettings().getValue('General:Character encoding, txt-files', 'cp1252')
#            out = codecs.open(fileName, mode = 'w', encoding = txtencode)
#            separator = '\t' # Use tab as item separator.
#            # Header.
#            out.write(separator.join(map(unicode, self._header)))
#            # Rows.
#            for row in self._rows:
#                # Use tab as column separator and CR/LF as row delimiter.
#                out.write(separator.join(map(unicode, row))')
#        except (IOError, OSError):
#            envmonlib.Logging().log("Failed to write to file: " + fileName)
#            raise
#        finally:
#            if out: out.close()

#    def SaveExcelXlsxFile(self, fileName):
#        """ """
#        if fileName == None:
#            raise UserWarning('File name is missing.')
#        try:
#            workbook = excelworkbook.Workbook() # Create workbook.
#            worksheet = workbook.get_active_sheet() # Workbooks contains at least one worksheet.
#            # Header.
#            for columnindex, item in enumerate(self._header):
#                cell = worksheet.cell(row=0, column=columnindex)
#                cell.value = unicode(item)
#            # Rows.
#            for rowindex, row in enumerate(self._rows):
#                for columnindex, item in enumerate(row):
#                    cell = worksheet.cell(row=rowindex + 1, column=columnindex)
#                    cell.value = unicode(item)
#            #    
#            excelwriter.save_workbook(workbook, fileName)
#        except (IOError, OSError):
#            envmonlib.Logging().log("Failed to write to file: " + fileName)
#            raise

#class TextFile(MonitoringFiles):
#    """ 
#    """
#    def __init__(self):
#        """ """
#        # Initialize parent.
#        super(TextFile, self).__init__()
#
#    def readFile(self, fileName = None, separator = '\t'):
#        """ """
#        if fileName == None:
#            raise UserWarning('File name is missing.')
#        #
#        file = None
#        try:
#            #
#            txtencode = toolbox_settings.ToolboxSettings().getValue('General:Character encoding, txt-files', 'cp1252')
#            file = codecs.open(fileName, mode = 'r', encoding = txtencode)
#            #
#            for index, row in enumerate(file):
#                if index == 0: # First row is assumed to be the header row.
#                    for headeritem in map(string.strip, row.split(separator)):
#                        self._header.append(headeritem)
#                else:
#                    newrow = []
#                    for item in map(string.strip, row.split(separator)):
#                        newrow.append(item)
#                    self._rows.append(newrow)
#        #  
#        except (IOError, OSError):
#            envmonlib.Logging().log("Can't read text file.")
#            raise
#        finally:
#            if file: file.close() 


#class ExcelXlsxFile(MonitoringFiles):
#    """ 
#    """
#    def __init__(self):
#        """ """
#        # Initialize parent.
#        super(ExcelXlsxFile, self).__init__()
#
#    def readFile(self, fileName = None, separator = '\t'):
#        """ """
#        if fileName == None:
#            raise UserWarning('File name is missing.')
#        #
#        try:
#            workbook = excelreader.load_workbook(fileName)
####            print(workbook.get_sheet_names())
#            worksheet = workbook.get_active_sheet()
####            cell = worksheet.cell('B3')
####            print(cell)
####            cell_value = worksheet.cell('B3').value
####            print(cell_value)
#            for row in xrange(0, worksheet.get_highest_row()): 
#                if row == 0: # First row is assumed to be the header row.
#                    for col in xrange(0, worksheet.get_highest_column()):
#                        value = worksheet.cell(row=row, column=col).value
#                        if value:
#                            self._header.append(unicode(value))
#                        else:
#                            self._header.append(u'')
#                else:
#                    newrow = []
#                    for col in xrange(0, worksheet.get_highest_column()):
#                        value = worksheet.cell(row=row, column=col).value
#                        if value:
#                            newrow.append(unicode(value))
#                        else:
#                            newrow.append(u'')
#                    self._rows.append(newrow)        
#        #  
#        except Exception:
#            envmonlib.Logging().log("Can't read Excel (.xlsx) file. The python package openpyxl is needed to import Excel files.")
#            raise


class SharkwebDownload(MonitoringFiles):
    """ 
    """
    def __init__(self):
        """ """
        # Initialize parent.
        super(SharkwebDownload, self).__init__()

    def getData(self, parameters = None):
        """ """
        self.clear()
        if parameters == None:
            raise UserWarning('Parameters are missing.')
        # URL and parameters. Use unicode and utf-8 to handle swedish characters.



#    http://produkter.smhi.se/sharkweb/shark_save.php?
#    action=download_sample
#    bounds=
#    year_from=2010
#    year_to=2010
#    month=06
#    datatype=Phytoplankton
#    parameter=
#    project_code=
#    orderer=
#    station_name=
#    taxon_name=
#    sample_table_view=sample_col_std
#    delimiters=point-tab
#    lineend=unix
#    headerlang=sv
        
        
#        url = u'http://test.mellifica.org/sharkweb/shark_php.php'
#        url = u'http://produkter.smhi.se/sharkweb/shark_save.php'
        url = u'http://sharkweb.smhi.se/shark_save.php'
        
##        parameters = {}
        parameters['action']=u'download_sample'
##        parameters['year_from']=u'2010'
##        parameters['year_to']=u'2010'
##        parameters['month']=u'06'
##        parameters['datatype']=u'Phytoplankton'
##        parameters['parameter']=u'CONC'
        parameters['sample_table_view']=u'sample_col_bio'
        parameters['delimiters']=u'point-tab'
        parameters['lineend']=u'unix'
        parameters['headerlang']=u'en'     
        
        
#        parameters['year_from']= u'1900'
#        parameters['year_to']= u'2010'
#        parameters['month']= ''
#        parameters['parameter']= ''
        
        
           
        
        parameters = dict([k, v.encode('utf-8')] for k, v in parameters.items())
        params = urllib.urlencode(parameters)
        #
        envmonlib.Logging().log('DEBUG: URL: ' + url)
        envmonlib.Logging().log('DEBUG: Parameters: ' + params)
        #   
        f = urllib.urlopen(url, params)
        for row, line in enumerate(f.readlines()):
            # Use this if data is delivered in rows with tab separated fields.
            if row == 0:
                self._header = map(string.strip, unicode(line).split('\t'))
            else:
                self._rows.append(map(string.strip, unicode(line).split('\t')))
#            # Use this if data is delivered in the Json format.
#            self._dataset = json.loads(line, encoding = encode)
        envmonlib.Logging().log('Received rows: ' + str(len(self._rows)))


class PwCsv(MonitoringFiles):
    """ """
    def __init__(self):
        """ """
        self._sample_info = {} # Information related to sample.
        self._aggregated_rows = [] # Pre-calculated aggregations of data.
        # Initialize parent.
        super(PwCsv, self).__init__()

#    def clear(self):
#        """ """
#        self._sample_info = {}
#        self._aggregated_rows = []
#        # Initialize parent.
#        super(PwCsv, self).clear()
#
#    def getSampleInfo(self):
#        """ """
#        return self._sample_info
#
#    def readFile(self, fileName = None):
#        """ """
#        if fileName == None:
#            raise UserWarning('File name is missing.')
#        file = None
#        try:
#            txtencode = toolbox_settings.ToolboxSettings().getValue('General:Character encoding, txt-files', 'cp1252')
#            file = codecs.open(fileName, mode = 'r', encoding = txtencode)
#            separator = ',' # Use ',' as item separator.
#            
#            # Read data header. Same header used for data and aggregated data.
#            self._header = []
#            for headeritem in file.readline().split(separator):
#                item = headeritem.strip().strip('"').strip()
#                self._header.append(item)
#                
#            # Empty line.
#            file.readline()
#            
#            # Read data rows. Continue until empty line occurs.
#            self._rows = []
#            row = file.readline()
#            while len(row.strip()) > 0:
#                rowitems = []
#                for item in row.split(separator):
#                    rowitems.append(item.strip().strip('"').strip())
#                self._rows.append(rowitems) 
#                row = file.readline()
#            
#            # Read aggregated data rows. Continue until empty line occurs.
#            self._aggregated_rows = []
#            row = file.readline()
#            while len(row.strip()) > 0:
#                rowitems = []
#                for item in row.split(separator):
#                    rowitems.append(item.strip().strip('"').strip())
#                self._aggregated_rows.append(rowitems) 
#                row = file.readline()
#                         
#            # Read total counted.
#            row = file.readline() # Not used...TODO:
#            row = file.readline() # Empty.
#            
#            # Read total counted.
#            row = file.readline() # Not used...TODO:
#            row = file.readline() # Empty.
#            
#            # Read chamber and magnification info.
##            self._aggregated_data['rows'] = []
#            row = file.readline()
#            while len(row.strip()) > 0:
#                rowitems = []
#                for item in row.split(separator):
#                    rowitems.append(item.strip().strip('"').strip())
##                self._aggregated_data['rows'].append(rowitems) # Not used...TODO: 
#                row = file.readline()
#                         
#            # Read info related to sample.
#            row = file.readline()
#            while len(row.strip()) > 0:
#                key, value = row.split(separator)
#                self._sample_info[key.strip().strip('"').strip()] = value.strip().strip('"').strip()
#                row = file.readline()
#                
#        except (IOError, OSError):
#            raise
#        finally:
#            if file: file.close()
#
#    def exportAsJson(self, file = None):
#        """ """
#        if file == None:
#            raise UserWarning('File name is missing.')
#        #
#        jsonencode = toolbox_settings.ToolboxSettings().getValue('General:Character encoding, json-files', 'cp1252')
#        outdata = open(file, 'w')
#        
#        jsonexport = {}
#        jsonexport['metadata'] = self._metadata
#        jsonexport['sample_info'] = self._sample_info
#        jsonexport['header'] = self._rows
#        jsonexport['rows'] = self._rows
#        jsonexport['aggregated_rows'] = self._aggregated_rows
#        outdata.write(json.dumps(jsonexport, encoding = jsonencode, 
#                                 sort_keys=True, indent=4))
#        outdata.close()


class PwCsvTable(MonitoringFiles):
    """
    PW data organised in table format. Sample data are added to each row.
    It is possible to aggregate data from multiple PW files. 
    """
    def __init__(self):
        """ """
        # Initialize parent.
        super(PwCsvTable, self).__init__()

#    def readFile(self, fileName = None):
#        """
#            # Keywords in the PW sample dictionary used for sample info: 
#            #    'Sample Id', 'Counted on', 'Chamber diam.', 
#            #    'Sampler', 'Latitude', 'StatName', 'Sample by', 'Date', 
#            #    'Sedim. time (hr)', 'No. Depths', 'Counted by', 
#            #    'Max. Depth', 'Longitude', 'Project', 'Depth', 
#            #    'Min. Depth', 'Time', 'Mixed volume', 'StatNo', 
#            #    'Comment', 'Sample size', 'Amt. preservative', 
#            #    'Sedim. volume', 'Ship', 'Preservative'
#        """
#        # Read the PW file into dictionaries and rows.
#        pwsample = PwCsv()
#        pwsample.readFile(fileName)
#        # Define column order and translations.
#        sampleinfoheaders = [
#            ['Sample Id', 'Sample id'], 
#            ['Project', 'Project'], 
#            ['Ship', 'Ship'],            
#            ['StatNo', 'Stat no'], 
#            ['StatName', 'Stat name'],              
#            ['Latitude', 'Latitude'], 
#            ['Longitude', 'Longitude'], 
#            ['Date', 'Date'], 
#            ['Time', 'Time'], 
#            ['Depth', 'Depth'], 
#            ['Min. Depth', 'Min. depth'], 
#            ['Max. Depth', 'Max. depth'], 
#            ['No. Depths', 'No. depths'],             
#            ['Sampler', 'Sampler'], 
#            ['Sample size', 'Sample size'], 
#            ['Sample by', 'Sample by'], 
#            ['Mixed volume', 'Mixed volume'], 
#            ['Amt. preservative', 'Amt. preservative'], 
#            ['Sedim. volume', 'Sedim. volume'], 
#            ['Preservative', 'Preservative'],
#            ['Sedim. time (hr)', 'Sedim. time (hr)'], 
#            ['Chamber diam.', 'Chamber diameter'],             
#            ['Counted on', 'Counted on'],
#            ['Counted by', 'Counted by'], 
#            ['Comment', 'Comment']
#        ]
#        # Header.
#        self._header = [] # Rewrite if multiple files are read.
#        for headerpair in sampleinfoheaders:
#                self._header.append(headerpair[1]) # Second part for header.
#        self._header = self._header + pwsample.getHeader()
#        # Data.
#        sampleinfo = pwsample.getSampleInfo()
#        for row in pwsample.getRows():
#            newrow = []
#            index = 0
#            # Get data for the sample info part.
#            for headerpair in sampleinfoheaders:
#                newrow.append(sampleinfo[headerpair[0]]) # First part as key.          
#            # Get table data. Use the same header order as in the PW file.
#            for item in row:
#                newrow.append(item)
#                index += 1
#            #
#            self._rows.append(newrow)

