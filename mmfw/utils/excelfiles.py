#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
# Project: Moray
# Author: Arnold Andreasson, info@mellifica.se
# Copyright (c) 2011 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License as follows:
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import mmfw
try: 
    import openpyxl.workbook as excelworkbook
    import openpyxl.reader.excel as excelreader
    import openpyxl.writer.excel as excelwriter
except ImportError: 
    print("Python package openpyxl missing. Download from http://pypi.python.org/pypi/openpyxl.")
    raise UserWarning("Python package openpyxl missing. Download from http://pypi.python.org/pypi/openpyxl.")

class ExcelFileReader():
    """
    This class ...  
    """
    def __init__(self):
        """ """
        self._metadata = {}
        self._header = [] # 
        self._rows = []
        
    def readFile(self, 
                 file_name = None,
                 sheet_name = None, 
                 header_row = 0, 
                 data_rows_from = 1, 
                 data_rows_to = None, # None = read all.
                 used_columns_from = 0, 
                 used_columns_to = None): # None = read all.
        """ """
        if file_name == None:
            raise UserWarning('File name is missing.')
        #
        try:
            workbook = excelreader.load_workbook(file_name)
            if workbook == None:
                raise UserWarning("Can't read Excel (.xlsx) file.")
            worksheet = None
            if sheet_name:
                # 
                if sheet_name in workbook.get_sheet_names():
                    worksheet = workbook.get_sheet_by_name(sheet_name)
                else:
                    raise UserWarning("Excel sheet " + sheet_name + " not available.")      
            else:
                # Use the first sheet if not specified.
                worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
            
            # Prepare boundary. 
            if used_columns_to:
                used_columns_to = min(used_columns_to,  worksheet.get_highest_column())
            else:
                used_columns_to = worksheet.get_highest_column()
            if data_rows_to:
                data_rows_to = min(data_rows_to,  worksheet.get_highest_row())
            else:
                data_rows_to = worksheet.get_highest_row()
                
            # Read header row.
            for col in xrange(used_columns_from, worksheet.get_highest_column()):
                value = worksheet.cell(row=header_row, column=col).value
                if value:
                    self._header.append(unicode(value).strip())
                else:
                    self._header.append(u'')
                    
            # Read data rows.
            for row in xrange(data_rows_from, data_rows_to): 
                newrow = []
                for col in xrange(used_columns_from, worksheet.get_highest_column()):
                    value = worksheet.cell(row=row, column=col).value
                    if value:
                        newrow.append(unicode(value).strip())
                    else:
                        newrow.append(u'')
                self._rows.append(newrow)
            # Close.
            # TODO: Not needed?        
        #  
        except Exception:
            print("Can't read Excel (.xlsx) file. File name: " + file_name )


    def clear(self):
        """ """
        self._metadata = {}
        self._header = []
        self._rows = []

    def getMetadata(self):
        """ """
        return self._metadata

    def getHeader(self):
        """ """
        return self._header

    def getRows(self):
        """ """
        return self._rows

    def getHeaderCell(self, column):
        """ """
        try:
            return self._header[column]
        except Exception:
            return ''

    def getColumnCount(self):
        """ """
        try:
            return len(self._header)
        except Exception:
            return 0

    def getRowCount(self):
        """"""
        try:
            return len(self._rows)
        except Exception:
            return 0

    def getDataCell(self, row, column):
        """ """
        try:
            return self._rows[row][column]
        except Exception:
            return ''

    def getDataCellByName(self, row, column_name):
        """ """
        try:
            column = self._header.index(column_name)
            return self._rows[row][column]
        except Exception:
            return ''
        
class ExcelFileWriter():
    """
    This class ...  
    """
    def __init__(self):
        """ """
        
    def SaveExcelXlsxFile(self, fileName):
        """ """
        if fileName == None:
            raise UserWarning('File name is missing.')
        try:
            workbook = excelworkbook.Workbook() # Create workbook.
            worksheet = workbook.get_active_sheet() # Workbooks contains at least one worksheet.
            # Header.
            for columnindex, item in enumerate(self._header):
                cell = worksheet.cell(row=0, column=columnindex)
                cell.value = unicode(item)
            # Rows.
            for rowindex, row in enumerate(self._rows):
                for columnindex, item in enumerate(row):
                    cell = worksheet.cell(row=rowindex + 1, column=columnindex)
                    cell.value = unicode(item)
            #    
            excelwriter.save_workbook(workbook, fileName)
            # Close.
            # TODO: Not needed?        
        except (IOError, OSError):
            mmfw.Logging().log("Failed to write to file: " + fileName)
            raise

