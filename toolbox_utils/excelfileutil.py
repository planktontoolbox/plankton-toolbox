#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
# Copyright (c) 2010-2015 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).
#
from __future__ import unicode_literals

import os 

try: 
    openpyxl_installed = False
    import openpyxl
    openpyxl_installed = True
except ImportError:
    print('Python package "openpyxl" is missing. Please install.')

class ExcelFileUtil():
    """ """
    def __init__(self):
        """ """
        self._metadata = {}
        self._header = [] # 
        self._rows = []
        
    def read_table_data(self, 
                        file_path,
                        file_name, # May contain both path and name if 'file_path' is empty.
                        sheet_name = None, 
                        header_row = 0, # If header not the first row. 
                        data_rows_from = 1, # If data not starts at second row. 
                        data_rows_to = None, # None = read all.
                        select_columns_by_name = [], # To reduce the number of columns.
                        select_columns_by_index = [], # To reduce the number of columns.
                        ):
        """ """
        table_header = [] 
        table_rows = []
        #
        if openpyxl_installed == False:
            raise UserWarning('Can\'t read .xlsx files ("openpyxl" is not installed).')
        # File path and name.
        filename = file_name
        if file_path and file_name:
            filename = os.path.join(file_path, file_name)
        if filename is None:
            raise UserWarning('File name is missing.')
        if not os.path.exists(filename):
            raise UserWarning('File is not found.  File: ' + filename)
        # 
        columnsbyindex = None
        if select_columns_by_index:
            columnsbyindex = select_columns_by_index
        #
        try:
            workbook = openpyxl.load_workbook(filename, use_iterators = True) # Supports big files.
            if workbook == None:
                raise UserWarning('Can\'t read Excel (.xlsx) file.')
            worksheet = None
            if sheet_name:
                # 
                if sheet_name in workbook.get_sheet_names():
                    worksheet = workbook.get_sheet_by_name(name = sheet_name)
                else:
                    raise UserWarning('Excel sheet ' + sheet_name + ' not available.')      
            else:
                # Use the first sheet if not specified.
                worksheet = workbook.get_sheet_by_name(name = workbook.get_sheet_names()[0])
            #
            header = []
            for rowindex, row in enumerate(worksheet.iter_rows()):
                if (data_rows_to is not None) and (rowindex > data_rows_to):
                    break # Break loop if data_row_to is defined and exceeded.
                elif rowindex == header_row:
                    for cell in row:
                        value = cell.value
                        if value == None:
                            header.append('')
                        else:
                            header.append(unicode(value).strip())
                    #
                    table_header = header
                elif rowindex >= data_rows_from:
                    newrow = []
                    for cell in row:
                        value = cell.value
                        if value == None:
                            newrow.append('')
                        else:
                            newrow.append(unicode(value).strip())
                    #
                    table_rows.append(newrow)
            #
            return (table_header, table_rows)
        #  
        except Exception as e:
            msg = 'Failed to read from file. File name: ' + filename + '. Exception: ' + unicode(e)
            print(msg)
            raise

    def write_table_data(self, 
                         file_path,
                         file_name, # May contain both path and name if 'file_path' is empty.
                         table_header,
                         table_rows
                         ):
        """ """
        if openpyxl_installed == False:
            raise UserWarning('Can\'t write to .xlsx files ("openpyxl" is not installed).')
        # File path and name.
        filename = file_name
        if file_path and file_name:
            filename = os.path.join(file_path, file_name)
        if filename is None:
            raise UserWarning('File name is missing.')
        if (file_path) and (not os.path.exists(file_path)):
            try:
                os.makedirs(file_path)
                print('Directories created for this path: ' + file_path)
            except Exception as e:
                raise UserWarning('Can\'t create directories in path. Path: ' + file_path + '. Exception: ' + e)
        try:
            workbook = openpyxl.Workbook(optimized_write = True)  # Supports big files.
            worksheet = workbook.create_sheet()
            # Header.
            worksheet.append(table_header)
            # Rows.
            for row in table_rows:
                worksheet.append(row)
            # Save to file.   
            workbook.save(file_name)
        #
        except Exception as e:
            msg = 'Failed to write to file: ' + file_name + '. Exception: ' + unicode(e)
            print(msg)
            raise
